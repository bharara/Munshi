from openpyxl import load_workbook
wb = load_workbook('Munshi.xlsx')

# Gets the Client List

client = wb.get_sheet_by_name('Client')
clist = []
for n in client['A']:
    if n.value == None:
        continue
    clist.append(n.value)
clist.remove('Name')

# Define a list of client names:

def num(n):

    # Make sure an input becomes an integer

    while True:
        try:
            int(n)
            return int(n)
        except:
            print("Enter a number please")
            n = input(" == ")
            continue

def in_choice(l,n):
    # Checks if a value is in a given list

    while True:
        if n in l:
            return n
        n=input("\nPlease Enter a valid value:")
        
def write_client(name):

    # Writes information of a client

    n = len(clist) + 2

    client['A'+str(n)] = name
    clist.append(name)

    print('\nName:', name)
    client['B'+str(n)] = input('Detail-1: ')
    client['C'+str(n)] = input('Detail-2: ')
    client['D'+str(n)] = input('Detail-3: ')
    client['E'+str(n)] = input('Detail-4: ')
    client['F'+str(n)] = input('Detail-5: ')

    print('\nNow to the entry:')
    print('Name:', name)

def is_client(name): # Varifies the name of the client

    while True:
        if name in clist:
            return name
        
        print('\nThis name have never been entered before. Is it a new client or you just mistyped?')
        print('I would re-enter the name (Press 1)')
        print("It's a new client I would like to add (Press 2)")
        ask = in_choice(['1','2'], input('=: '))
        
        if ask == '1':
            name = input('Name: ')
        elif ask == '2':
            write_client(name)

def get_input():

    # Input Module without checks and Graphics

    data = wb.get_sheet_by_name('Data')
    
    print("\nEnter 0 in name feild to save and return\nEnter -1 to return without saving")
    
    for n in range(data.max_row + 1,200):
        name = input('Name: ')
        if name == '0':
            print('\nThanks for saving ... \n')
            wb.save('Munshi.xlsx')
            break
        elif name == '-1':
            print('\nReturning without saving ... \n')
            break

        data['A'+str(n)] = is_client(name)
        data['B'+str(n)] = input('Date: ')
        data['C'+str(n)] = input('Detail-1: ')
        data['D'+str(n)] = input('Detail-2: ')
        data['E'+str(n)] = input('Detail-3: ')
        data['F'+str(n)] = num(input('Credit: '))
        data['G'+str(n)] = num(input('Debit: '))
        print()

# Main Program

while True:
    print("Hello and Welcome to accounts manager")
    print("Please chose a task to perform below:")
    print("Press 0 to exit")
    print("Press 1 to input data")
    action=in_choice(['0','1'],input('=: '))

    if action=='1': get_input()
    elif action=='0': break
    print()
