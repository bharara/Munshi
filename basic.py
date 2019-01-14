from flask import Flask, render_template, request, redirect
from openpyxl import load_workbook
import webbrowser
from datetime import datetime

app = Flask(__name__)

# Load workbooks and sheets
wb = load_workbook(filename='Munshi.xlsx')
data = wb.get_sheet_by_name('Data')
client = wb.get_sheet_by_name('Client')
roker = data['K1'].value

# Gets the client list
clist = []
for n in client['A']:
    if n.value == None:
        break
    clist.append(n.value)
clist.remove('Name')
# clist.sort()
num = len(clist) + 1

# Gets necessary variables
in_detail1 = data['C1'].value
in_detail2 = data['D1'].value
in_detail3 = data['E1'].value
c_detail1 = client['B1'].value
c_detail2 = client['C1'].value
c_detail3 = client['D1'].value
c_detail4 = client['E1'].value
c_detail5 = client['F1'].value
name1 = 'Name'

# Main app
@app.route('/')
def index():
	return render_template('home.html', roker=roker)

# Input Page
@app.route('/input')
def input():
	name = ""
	return render_template('input.html', clist=clist, detail1=in_detail1, detail2=in_detail2, detail3=in_detail3, name=name)

# Writes Input Data on file
@app.route('/input', methods=['POST'])
def input_after():

	wb = load_workbook(filename='Munshi.xlsx')
	data = wb.get_sheet_by_name('Data')
	
	# Count Already inputed data
	n=1
	for c in data['A']:
		if c.value == None:
			break
		n+=1

	# Writes data on file
	name = request.form["name"].title()
	credit = int(request.form['credit'])
	debit = int(request.form['debit'])
	data['A'+str(n)] = name
	data['B'+str(n)] = datetime.strptime(request.form['date'],"%Y-%m-%d").strftime('%d-%m-%y')
	data['C'+str(n)] = request.form['detail1']
	data['D'+str(n)] = request.form['detail2']
	data['E'+str(n)] = request.form['detail3']
	data['F'+str(n)] = credit
	data['G'+str(n)] = debit
	
	# Updates Roker in memory and file
	global roker
	roker = roker + credit - debit
	data['K1'] = roker
	wb.save('Munshi.xlsx')

	# For new clients call add_client functiom
	if name in clist:
		return render_template('input.html', clist=clist, detail1=in_detail1, detail2=in_detail2, detail3=in_detail3)
	else:
		# Return to input template with all arguments
		global name1
		name1 = name
		return redirect('/add_client')

# Client Addition App
@app.route('/add_client')
def add_client():
	return render_template('add_client.html',clist=clist, detail1=c_detail1, detail2=c_detail2, detail3=c_detail3, detail4=c_detail4, detail5=c_detail5, name=name1)

# Writes client data on file
@app.route('/add_client', methods=['POST'])
def add_clients_after():
	wb = load_workbook(filename='Munshi.xlsx')
	client = wb.get_sheet_by_name('Client')

	num=1
	for c in client['A']:
		if c.value == None:
			break
		num+=1

	# Adding client in worksheet
	client['A'+str(num)] = request.form["name"].title()
	clist.append(request.form["name"].title())
	client['B'+str(num)] = request.form["detail1"]
	client['C'+str(num)] = request.form["detail2"]
	client['D'+str(num)] = request.form["detail3"]
	client['E'+str(num)] = request.form["detail4"]
	client['F'+str(num)] = request.form["detail5"]
	wb.save('Munshi.xlsx')
	# Redirecting the user to add_client page
	return render_template('add_client.html',clist=clist, detail1=c_detail1, detail2=c_detail2, detail3=c_detail3, detail4=c_detail4, detail5=c_detail5)

# Report Main Page
@app.route('/report')
def report():
	return render_template('report.html')

# Client Report
@app.route('/report/user')
def user():
	return render_template('user.html', clist=sorted(clist))

# Show user report by form
@app.route('/report/user', methods=['POST'])
def user_after():
	name = request.form['name']
	return redirect('/report/user/'+name)

#User Report Shown
@app.route('/report/user/<name>')
def user2(name):

    wb = load_workbook(filename='Munshi.xlsx', data_only=True)
    client = wb.get_sheet_by_name('Client')
    cid = clist.index(name) + 2
 
    cd1 = client['B'+str(cid)].value
    cd2 = client['C'+str(cid)].value
    cd3 = client['D'+str(cid)].value
    cd4 = client['E'+str(cid)].value
    cd5 = client['F'+str(cid)].value
    credit = client['G'+str(cid)].value
    debit = client['H'+str(cid)].value

    name = name.replace('%20',' ')

    outlist= []
    n=1
    for val in data['A']:
        if val.value == name:
        	try:
        		hall = data['B'+str(n)].value.strftime("%d-%B-%Y")
        	except:
        		hall = datetime.strptime(data['B'+str(n)].value,"%d-%m-%y").strftime("%d-%B-%Y")
        	
        	outlist.append([hall, data['C'+str(n)].value, data['D'+str(n)].value, data['E'+str(n)].value, data['F'+str(n)].value, data['G'+str(n)].value])
        n+=1

    return render_template('user2.html', outlist=outlist, in_detail1=in_detail1, in_detail2 = in_detail2, in_detail3 = in_detail3,
                           c_detail1 = c_detail1, c_detail2 = c_detail2, c_detail3 = c_detail3, c_detail4 = c_detail4, c_detail5 = c_detail5,
                           cd1=cd1, cd2=cd2, cd3=cd3, cd4=cd4, cd5=cd5, credit=credit, debit=debit, name=name)

# Monthly Report
@app.route('/report/month')
def month():

	wb = load_workbook(filename='Munshi.xlsx', data_only=True)
	month = wb.get_sheet_by_name('Month')

	month_list = []

	for n in range(2,50):
		if month['D'+str(n)].value == 0:
			break
		else:
			sublist = []
			sublist.append(month['A'+str(n)].value.strftime("%B-%Y"))
			sublist.append(month['B'+str(n)].value)
			sublist.append(month['C'+str(n)].value)
			sublist.append(month['D'+str(n)].value)
			month_list.append(sublist)

	return render_template('month.html', month_list=month_list)

# Summary
@app.route('/report/summary')
def summary():
	wb = load_workbook(filename='Munshi.xlsx', data_only=True)
	client = wb.get_sheet_by_name('Client')

	n=2
	lst = []
	for val in client['A']:
		if client['C'+str(n)].value != '-':
			break
		lst.append([client['A'+str(n)].value, client['B'+str(n)].value, client['C'+str(n)].value, client['D'+str(n)].value, client['E'+str(n)].value, client['F'+str(n)].value, client['G'+str(n)].value, client['H'+str(n)].value])
		n+=1

	return render_template('summary.html', roker=roker, lst=lst, c_detail1 = c_detail1, c_detail2 = c_detail2, c_detail3 = c_detail3, c_detail4 = c_detail4, c_detail5 = c_detail5)

# Options main page
@app.route('/options')
def options():
	return render_template('options.html')

# Change Input Label
@app.route('/options/input')
def label_input():
	return render_template('label_input.html', detail1=in_detail1, detail2=in_detail2, detail3=in_detail3)

# Write Label Names
@app.route('/options/input', methods=['POST'])
def label_input_after():

	global in_detail1, in_detail2, in_detail3
	wb = load_workbook(filename='Munshi.xlsx')
	data = wb.get_sheet_by_name('Data')
	
	# Store them in Variables
	in_detail1 = request.form["detail1"]
	in_detail2 = request.form['detail2']
	in_detail3 = request.form['detail3']

	# Write Variables on File
	data['C1'] = in_detail1
	data['D1'] = in_detail2
	data['E1'] = in_detail3

	wb.save('Munshi.xlsx')

	return redirect('/options')

# Change Client Label
@app.route('/options/client')
def label_client():
	return render_template('label_client.html', detail1=c_detail1, detail2=c_detail2, detail3=c_detail3, detail4=c_detail4, detail5=c_detail5)

# Write Label Names
@app.route('/options/client', methods=['POST'])
def label_client_after():

	global c_detail1, c_detail2, c_detail3, c_detail4, c_detail5
	wb = load_workbook(filename='Munshi.xlsx')
	client = wb.get_sheet_by_name('Client')
	
	# Store them in Variables
	c_detail1 = request.form["detail1"]
	c_detail2 = request.form['detail2']
	c_detail3 = request.form['detail3']
	c_detail4 = request.form['detail4']
	c_detail5 = request.form['detail5']

	# Write Variables on File
	client['B1'] = c_detail1
	client['C1'] = c_detail2
	client['D1'] = c_detail3
	client['E1'] = c_detail4
	client['F1'] = c_detail5

	wb.save('Munshi.xlsx')

	return redirect('/options')

# About Page
@app.route('/about')
def about():
	return render_template('about.html')


# Autodebug and Run the app
if __name__ == '__main__':
	webbrowser.open('http://127.0.0.1:5000', new=2, autoraise=True)
	app.run(debug=True)
